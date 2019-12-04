import openpyxl

wb = openpyxl.load_workbook('/Users/davidguo/Documents/Pythom/automate_online-materials/produceSales.xlsx')


sheet = wb['Sheet']

pricedict = {
  "Celery": 1.19,
  "Garlic": 3.07,
  "Lemon": 1.27,
}

for i in range(2, sheet.max_row):
	if sheet.cell(row=i, column=1).value == 'Celery':
		sheet.cell(row=i, column=2).value = pricedict['Celery']
	elif sheet.cell(row=i, column=1).value == 'Garlic':
		sheet.cell(row=i, column=2).value = pricedict['Garlic']
	elif sheet.cell(row=i, column=1).value == 'Lemon':
		sheet.cell(row=i, column=2).value = pricedict['Lemon']

wb.save('updatedProduceSales.xlsx')
