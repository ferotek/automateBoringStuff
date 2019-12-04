import openpyxl
wb = openpyxl.load_workbook('automate_online-materials/censuspopdata.xlsx')
sheet = wb['Population by Census Tract']

dictcount = {}
dictpop2010 = {}

for i in range(2, sheet.max_row):
	if sheet.cell(row=i, column=3).value in dictcount:
		dictcount[sheet.cell(row=i, column=3).value] = dictcount[sheet.cell(row=i, column=3).value] + 1
	else:
		dictcount[sheet.cell(row=i, column=3).value] = 1

for i in range(2, sheet.max_row):
	if sheet.cell(row=i, column=3).value in dictpop2010:
		dictpop2010[sheet.cell(row=i, column=3).value] = dictpop2010[sheet.cell(row=i, column=3).value] + int(sheet.cell(row=i, column=4).value)
	else:
		dictpop2010[sheet.cell(row=i, column=3).value] = int(sheet.cell(row=i, column=4).value)

census = open('census.txt', 'w')

#print(county, dictcount[county], dictpop2010[county])

for county in dictcount:
	L = [county + ' ', str(dictcount[county])+ ' ', str(dictpop2010[county])+ '\n']
	census.writelines(L)
  